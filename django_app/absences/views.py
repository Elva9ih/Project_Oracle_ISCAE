import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.db import connection
from django.db.models import Count, Q

from .models import (
    Etudiant, Enseignant, Matiere, Seance, Absence,
    Justificatif, AuditLog, Notification, Groupe, Filiere, Semestre,
    UserProfile
)
from .forms import (
    EtudiantForm, EnseignantForm, SeanceForm,
    MarquerAbsenceForm, JustificatifForm, TraiterJustificatifForm,
    FiltreAbsencesForm, UserCreateForm
)
from .decorators import get_user_role, role_required, admin_required, enseignant_required
from . import oracle_services


# ============================================================
# DASHBOARD (all roles, different content)
# ============================================================

@login_required
def dashboard(request):
    role = get_user_role(request.user)

    if role == 'ETUDIANT':
        return dashboard_etudiant(request)
    elif role == 'ENSEIGNANT':
        return dashboard_enseignant(request)
    else:
        return dashboard_admin(request)


def dashboard_admin(request):
    try:
        kpi = oracle_services.get_kpi_global()
    except Exception:
        kpi = {}
    try:
        stats_groupe = oracle_services.get_stats_par_groupe()[:10]
    except Exception:
        stats_groupe = []
    try:
        evolution = oracle_services.get_evolution_mensuelle()
    except Exception:
        evolution = []

    context = {
        'kpi': kpi,
        'stats_groupe': stats_groupe,
        'evolution_labels': json.dumps([e.get('MOIS', '') for e in evolution]),
        'evolution_total': json.dumps([int(e.get('TOTAL_ABSENCES', 0)) for e in evolution]),
        'evolution_nj': json.dumps([int(e.get('NON_JUSTIFIEES', 0)) for e in evolution]),
        'evolution_j': json.dumps([int(e.get('JUSTIFIEES', 0)) for e in evolution]),
        'groupe_labels': json.dumps([g.get('CODE_GROUPE', '') for g in stats_groupe]),
        'groupe_absences': json.dumps([int(g.get('TOTAL_ABSENCES', 0)) for g in stats_groupe]),
    }
    return render(request, 'absences/dashboard.html', context)


def dashboard_enseignant(request):
    try:
        profile = request.user.profile
        enseignant = profile.id_enseignant
    except Exception:
        enseignant = None

    seances = []
    if enseignant:
        seances = Seance.objects.filter(id_enseignant=enseignant).select_related(
            'id_matiere', 'id_groupe'
        ).order_by('-date_seance')[:20]

    return render(request, 'absences/dashboard_enseignant.html', {
        'enseignant': enseignant,
        'seances': seances,
    })


def dashboard_etudiant(request):
    try:
        profile = request.user.profile
        etudiant = profile.id_etudiant
    except Exception:
        etudiant = None

    absences = []
    notifications = []
    nb_absences = 0
    nb_non_just = 0
    taux = 0
    statut_abs = 'N/A'

    if etudiant:
        absences = Absence.objects.filter(id_etudiant=etudiant).select_related(
            'id_seance', 'id_seance__id_matiere'
        ).order_by('-date_absence')[:20]
        notifications = Notification.objects.filter(id_etudiant=etudiant).order_by('-date_creation')[:10]
        try:
            nb_absences = oracle_services.get_nb_absences(etudiant.id_etudiant)
            nb_non_just = oracle_services.get_nb_absences_non_justifiees(etudiant.id_etudiant)
            taux = oracle_services.get_taux_absence(etudiant.id_etudiant)
            statut_abs = oracle_services.get_statut_absences(etudiant.id_etudiant)
        except Exception:
            nb_absences = absences.count()

    return render(request, 'absences/dashboard_etudiant.html', {
        'etudiant': etudiant,
        'absences': absences,
        'notifications': notifications,
        'nb_absences': nb_absences,
        'nb_non_just': nb_non_just,
        'taux': taux,
        'statut_abs': statut_abs,
    })


# ============================================================
# CRUD ÉTUDIANTS (admin only for create/update/delete)
# ============================================================

@login_required
def etudiant_list(request):
    role = get_user_role(request.user)
    if role == 'ETUDIANT':
        return redirect('dashboard')

    query = request.GET.get('q', '')
    groupe_id = request.GET.get('groupe', '')
    statut = request.GET.get('statut', '')

    etudiants = Etudiant.objects.all().select_related('id_groupe')

    # Enseignant: only see students in their groups
    if role == 'ENSEIGNANT':
        try:
            ens = request.user.profile.id_enseignant
            if ens:
                groupe_ids = Seance.objects.filter(id_enseignant=ens).values_list('id_groupe', flat=True).distinct()
                etudiants = etudiants.filter(id_groupe__in=groupe_ids)
        except Exception:
            pass

    if query:
        etudiants = etudiants.filter(
            Q(nom__icontains=query) | Q(prenom__icontains=query) | Q(cne__icontains=query)
        )
    if groupe_id:
        etudiants = etudiants.filter(id_groupe=groupe_id)
    if statut:
        etudiants = etudiants.filter(statut=statut)

    etudiants = etudiants.order_by('nom', 'prenom')
    paginator = Paginator(etudiants, 25)
    page = request.GET.get('page')
    etudiants_page = paginator.get_page(page)
    groupes = Groupe.objects.all().order_by('code_groupe')

    return render(request, 'absences/etudiant_list.html', {
        'etudiants': etudiants_page,
        'groupes': groupes,
        'query': query,
        'groupe_id': groupe_id,
        'statut': statut,
    })


@login_required
def etudiant_detail(request, pk):
    etudiant = get_object_or_404(Etudiant, pk=pk)

    # Students can only see their own detail
    role = get_user_role(request.user)
    if role == 'ETUDIANT':
        try:
            if request.user.profile.id_etudiant != etudiant:
                return redirect('dashboard')
        except Exception:
            return redirect('dashboard')

    absences = Absence.objects.filter(id_etudiant=etudiant).select_related('id_seance').order_by('-date_absence')[:20]

    try:
        nb_absences = oracle_services.get_nb_absences(pk)
        nb_non_just = oracle_services.get_nb_absences_non_justifiees(pk)
        taux = oracle_services.get_taux_absence(pk)
        statut_abs = oracle_services.get_statut_absences(pk)
    except Exception:
        nb_absences = absences.count()
        nb_non_just = absences.filter(est_justifiee=0).count()
        taux = 0
        statut_abs = 'N/A'

    notifications = Notification.objects.filter(id_etudiant=etudiant).order_by('-date_creation')[:10]

    return render(request, 'absences/etudiant_detail.html', {
        'etudiant': etudiant,
        'absences': absences,
        'nb_absences': nb_absences,
        'nb_non_just': nb_non_just,
        'taux': taux,
        'statut_abs': statut_abs,
        'notifications': notifications,
    })


@login_required
@admin_required
def etudiant_create(request):
    if request.method == 'POST':
        form = EtudiantForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Étudiant créé avec succès.')
            return redirect('etudiant_list')
    else:
        form = EtudiantForm()
    return render(request, 'absences/form_generic.html', {'form': form, 'title': 'Ajouter un étudiant'})


@login_required
@admin_required
def etudiant_update(request, pk):
    etudiant = get_object_or_404(Etudiant, pk=pk)
    if request.method == 'POST':
        form = EtudiantForm(request.POST, instance=etudiant)
        if form.is_valid():
            form.save()
            messages.success(request, 'Étudiant modifié avec succès.')
            return redirect('etudiant_detail', pk=pk)
    else:
        form = EtudiantForm(instance=etudiant)
    return render(request, 'absences/form_generic.html', {'form': form, 'title': 'Modifier l\'étudiant'})


@login_required
@admin_required
def etudiant_delete(request, pk):
    etudiant = get_object_or_404(Etudiant, pk=pk)
    if request.method == 'POST':
        etudiant.delete()
        messages.success(request, 'Étudiant supprimé.')
        return redirect('etudiant_list')
    return render(request, 'absences/confirm_delete.html', {'object': etudiant, 'title': 'Supprimer l\'étudiant'})


# ============================================================
# CRUD ENSEIGNANTS (admin only)
# ============================================================

@login_required
@admin_required
def enseignant_list(request):
    enseignants = Enseignant.objects.all().order_by('nom', 'prenom')
    paginator = Paginator(enseignants, 25)
    page = request.GET.get('page')
    return render(request, 'absences/enseignant_list.html', {'enseignants': paginator.get_page(page)})


@login_required
@admin_required
def enseignant_create(request):
    if request.method == 'POST':
        form = EnseignantForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Enseignant créé avec succès.')
            return redirect('enseignant_list')
    else:
        form = EnseignantForm()
    return render(request, 'absences/form_generic.html', {'form': form, 'title': 'Ajouter un enseignant'})


@login_required
@admin_required
def enseignant_update(request, pk):
    enseignant = get_object_or_404(Enseignant, pk=pk)
    if request.method == 'POST':
        form = EnseignantForm(request.POST, instance=enseignant)
        if form.is_valid():
            form.save()
            messages.success(request, 'Enseignant modifié.')
            return redirect('enseignant_list')
    else:
        form = EnseignantForm(instance=enseignant)
    return render(request, 'absences/form_generic.html', {'form': form, 'title': 'Modifier l\'enseignant'})


# ============================================================
# SÉANCES (admin + enseignant)
# ============================================================

@login_required
@enseignant_required
def seance_list(request):
    role = get_user_role(request.user)
    seances = Seance.objects.all().select_related('id_matiere', 'id_enseignant', 'id_groupe')

    if role == 'ENSEIGNANT':
        try:
            ens = request.user.profile.id_enseignant
            if ens:
                seances = seances.filter(id_enseignant=ens)
        except Exception:
            pass

    seances = seances.order_by('-date_seance')
    paginator = Paginator(seances, 25)
    page = request.GET.get('page')
    return render(request, 'absences/seance_list.html', {'seances': paginator.get_page(page)})


@login_required
@admin_required
def seance_create(request):
    if request.method == 'POST':
        form = SeanceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Séance créée.')
            return redirect('seance_list')
    else:
        form = SeanceForm()
    return render(request, 'absences/form_generic.html', {'form': form, 'title': 'Ajouter une séance'})


# ============================================================
# GESTION DES ABSENCES (admin + enseignant)
# ============================================================

@login_required
def absence_list(request):
    role = get_user_role(request.user)
    form = FiltreAbsencesForm(request.GET)
    absences = Absence.objects.all().select_related(
        'id_etudiant', 'id_seance', 'id_seance__id_matiere'
    ).order_by('-date_absence')

    # Etudiant: only their own
    if role == 'ETUDIANT':
        try:
            etud = request.user.profile.id_etudiant
            if etud:
                absences = absences.filter(id_etudiant=etud)
        except Exception:
            absences = absences.none()

    # Enseignant: only their groups
    elif role == 'ENSEIGNANT':
        try:
            ens = request.user.profile.id_enseignant
            if ens:
                absences = absences.filter(id_seance__id_enseignant=ens)
        except Exception:
            pass

    if form.is_valid():
        if form.cleaned_data.get('groupe'):
            absences = absences.filter(id_etudiant__id_groupe=form.cleaned_data['groupe'])
        if form.cleaned_data.get('matiere'):
            absences = absences.filter(id_seance__id_matiere=form.cleaned_data['matiere'])
        if form.cleaned_data.get('date_debut'):
            absences = absences.filter(date_absence__gte=form.cleaned_data['date_debut'])
        if form.cleaned_data.get('date_fin'):
            absences = absences.filter(date_absence__lte=form.cleaned_data['date_fin'])
        if form.cleaned_data.get('justifiee'):
            absences = absences.filter(est_justifiee=int(form.cleaned_data['justifiee']))

    paginator = Paginator(absences, 30)
    page = request.GET.get('page')

    return render(request, 'absences/absence_list.html', {
        'absences': paginator.get_page(page),
        'form': form,
    })


@login_required
@enseignant_required
def marquer_absences(request):
    groupe_id = request.GET.get('groupe') or request.POST.get('groupe')

    if request.method == 'POST' and 'etudiants' in request.POST:
        seance_id = request.POST.get('seance')
        etudiant_ids = request.POST.getlist('etudiants')
        saisi_par = request.user.username

        try:
            oracle_services.marquer_absences_bulk(
                int(seance_id),
                [int(eid) for eid in etudiant_ids],
                saisi_par
            )
            messages.success(request, f'{len(etudiant_ids)} absence(s) enregistrée(s) avec succès.')
            return redirect('absence_list')
        except Exception as e:
            messages.error(request, f'Erreur Oracle: {e}')

    groupes = Groupe.objects.all().order_by('code_groupe')

    # Enseignant: only their groups
    role = get_user_role(request.user)
    if role == 'ENSEIGNANT':
        try:
            ens = request.user.profile.id_enseignant
            if ens:
                groupe_ids = Seance.objects.filter(id_enseignant=ens).values_list('id_groupe', flat=True).distinct()
                groupes = groupes.filter(id_groupe__in=groupe_ids)
        except Exception:
            pass

    form = MarquerAbsenceForm(groupe_id=groupe_id) if groupe_id else None

    return render(request, 'absences/marquer_absences.html', {
        'form': form,
        'groupes': groupes,
        'groupe_id': groupe_id,
    })


# ============================================================
# JUSTIFICATIFS
# ============================================================

@login_required
def justificatif_list(request):
    role = get_user_role(request.user)
    justificatifs = Justificatif.objects.all().select_related(
        'id_absence', 'id_absence__id_etudiant', 'id_absence__id_seance'
    ).order_by('-date_soumission')

    if role == 'ETUDIANT':
        try:
            etud = request.user.profile.id_etudiant
            if etud:
                justificatifs = justificatifs.filter(id_absence__id_etudiant=etud)
        except Exception:
            justificatifs = justificatifs.none()

    statut = request.GET.get('statut', '')
    if statut:
        justificatifs = justificatifs.filter(statut=statut)

    paginator = Paginator(justificatifs, 25)
    page = request.GET.get('page')
    return render(request, 'absences/justificatif_list.html', {
        'justificatifs': paginator.get_page(page),
        'statut': statut,
    })


@login_required
def justificatif_create(request, absence_id):
    absence = get_object_or_404(Absence, pk=absence_id)

    # Students can only justify their own absences
    role = get_user_role(request.user)
    if role == 'ETUDIANT':
        try:
            if request.user.profile.id_etudiant != absence.id_etudiant:
                messages.error(request, "Vous ne pouvez justifier que vos propres absences.")
                return redirect('dashboard')
        except Exception:
            return redirect('dashboard')

    if request.method == 'POST':
        form = JustificatifForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                fichier_path = None
                if form.cleaned_data.get('fichier'):
                    f = form.cleaned_data['fichier']
                    fichier_path = f'justificatifs/{f.name}'
                    with open(f'media/{fichier_path}', 'wb+') as dest:
                        for chunk in f.chunks():
                            dest.write(chunk)

                oracle_services.justifier_absence(
                    absence_id,
                    form.cleaned_data['type_justificatif'],
                    form.cleaned_data['description'],
                    fichier_path
                )
                messages.success(request, 'Justificatif soumis avec succès.')
                return redirect('absence_list')
            except Exception as e:
                messages.error(request, f'Erreur: {e}')
    else:
        form = JustificatifForm()

    return render(request, 'absences/form_generic.html', {
        'form': form,
        'title': f'Justifier l\'absence #{absence_id}',
    })


@login_required
@admin_required
def justificatif_traiter(request, pk):
    justificatif = get_object_or_404(Justificatif, pk=pk)
    absence = justificatif.id_absence
    etudiant = absence.id_etudiant
    seance = absence.id_seance

    if request.method == 'POST':
        form = TraiterJustificatifForm(request.POST)
        if form.is_valid():
            try:
                oracle_services.traiter_justificatif(
                    pk,
                    form.cleaned_data['statut'],
                    form.cleaned_data.get('commentaire'),
                    request.user.username
                )
                messages.success(request, 'Justificatif traité.')
                return redirect('justificatif_list')
            except Exception as e:
                messages.error(request, f'Erreur: {e}')
    else:
        form = TraiterJustificatifForm()

    return render(request, 'absences/justificatif_traiter.html', {
        'form': form,
        'justificatif': justificatif,
        'absence': absence,
        'etudiant': etudiant,
        'seance': seance,
    })


# ============================================================
# STATISTIQUES (admin + enseignant)
# ============================================================

@login_required
@enseignant_required
def stats_par_groupe(request):
    try:
        stats = oracle_services.get_stats_par_groupe()
    except Exception:
        stats = []
    return render(request, 'absences/stats_groupe.html', {'stats': stats})


@login_required
@enseignant_required
def stats_par_matiere(request):
    try:
        stats = oracle_services.get_stats_par_matiere()
    except Exception:
        stats = []
    return render(request, 'absences/stats_matiere.html', {'stats': stats})


# ============================================================
# AUDIT LOG (admin only)
# ============================================================

@login_required
@admin_required
def audit_log(request):
    logs = AuditLog.objects.all().order_by('-date_operation')
    table = request.GET.get('table', '')
    operation = request.GET.get('operation', '')
    if table:
        logs = logs.filter(table_name=table)
    if operation:
        logs = logs.filter(operation=operation)

    paginator = Paginator(logs, 30)
    page = request.GET.get('page')
    return render(request, 'absences/audit_log.html', {
        'logs': paginator.get_page(page),
        'table': table,
        'operation': operation,
    })


# ============================================================
# TRAITEMENTS ORACLE (admin only)
# ============================================================

@login_required
@admin_required
def traitements(request):
    semestres = Semestre.objects.all()

    if request.method == 'POST':
        action = request.POST.get('action')
        try:
            if action == 'refresh_mv':
                oracle_services.refresh_materialized_views()
                messages.success(request, 'Vues matérialisées rafraîchies avec succès.')
            elif action == 'cloturer':
                sem_id = request.POST.get('semestre')
                oracle_services.cloturer_semestre(int(sem_id), request.user.username)
                messages.success(request, 'Semestre clôturé avec succès.')
        except Exception as e:
            messages.error(request, f'Erreur Oracle: {e}')
        return redirect('traitements')

    return render(request, 'absences/traitements.html', {'semestres': semestres})


# ============================================================
# USER MANAGEMENT (admin only)
# ============================================================

@login_required
@admin_required
def user_list(request):
    users = User.objects.all().select_related('profile').order_by('username')
    return render(request, 'absences/user_list.html', {'users': users})


@login_required
@admin_required
def user_create(request):
    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        if form.is_valid():
            user = User.objects.create_user(
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password'],
                email=form.cleaned_data.get('email', ''),
                first_name=form.cleaned_data.get('first_name', ''),
                last_name=form.cleaned_data.get('last_name', ''),
            )
            profile = UserProfile.objects.create(
                user=user,
                role=form.cleaned_data['role'],
                id_etudiant=form.cleaned_data.get('id_etudiant'),
                id_enseignant=form.cleaned_data.get('id_enseignant'),
            )
            messages.success(request, f'Utilisateur {user.username} créé avec le rôle {profile.get_role_display()}.')
            return redirect('user_list')
    else:
        form = UserCreateForm()
    return render(request, 'absences/form_generic.html', {'form': form, 'title': 'Créer un utilisateur'})


@login_required
@admin_required
def user_delete(request, pk):
    user = get_object_or_404(User, pk=pk)
    if user == request.user:
        messages.error(request, "Vous ne pouvez pas supprimer votre propre compte.")
        return redirect('user_list')
    if request.method == 'POST':
        user.delete()
        messages.success(request, 'Utilisateur supprimé.')
        return redirect('user_list')
    return render(request, 'absences/confirm_delete.html', {'object': user, 'title': 'Supprimer l\'utilisateur'})


# ============================================================
# EXPORTS PDF / EXCEL (admin + enseignant)
# ============================================================

@login_required
@enseignant_required
def export_absences_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Absences"

    headers = ['CNE', 'Nom', 'Prénom', 'Groupe', 'Matière', 'Date', 'Heure', 'Justifiée', 'Motif']
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Use raw SQL to avoid oracledb type issues
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT e.cne, e.nom, e.prenom, g.code_groupe, m.nom_matiere,
                   TO_CHAR(a.date_absence, 'YYYY-MM-DD'),
                   s.heure_debut || '-' || s.heure_fin,
                   CASE WHEN a.est_justifiee = 1 THEN 'Oui' ELSE 'Non' END,
                   a.motif
            FROM ABSENCES a
            JOIN ETUDIANTS e ON a.id_etudiant = e.id_etudiant
            JOIN GROUPES g ON e.id_groupe = g.id_groupe
            JOIN SEANCES s ON a.id_seance = s.id_seance
            JOIN MATIERES m ON s.id_matiere = m.id_matiere
            ORDER BY a.date_absence DESC
            FETCH FIRST 5000 ROWS ONLY
        """)
        for row_idx, row in enumerate(cursor.fetchall(), 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(val) if val else '')

    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="absences.xlsx"'
    wb.save(response)
    return response


@login_required
@enseignant_required
def export_absences_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from io import BytesIO

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Rapport des Absences - ISCAE", styles['Title']))
    elements.append(Spacer(1, 20))

    data = [['CNE', 'Nom', 'Prénom', 'Groupe', 'Matière', 'Date', 'Justifiée']]

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT e.cne, e.nom, e.prenom, g.code_groupe, m.nom_matiere,
                   TO_CHAR(a.date_absence, 'YYYY-MM-DD'),
                   CASE WHEN a.est_justifiee = 1 THEN 'Oui' ELSE 'Non' END
            FROM ABSENCES a
            JOIN ETUDIANTS e ON a.id_etudiant = e.id_etudiant
            JOIN GROUPES g ON e.id_groupe = g.id_groupe
            JOIN SEANCES s ON a.id_seance = s.id_seance
            JOIN MATIERES m ON s.id_matiere = m.id_matiere
            ORDER BY a.date_absence DESC
            FETCH FIRST 500 ROWS ONLY
        """)
        for row in cursor.fetchall():
            data.append([str(v) if v else '' for v in row])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="absences.pdf"'
    return response
